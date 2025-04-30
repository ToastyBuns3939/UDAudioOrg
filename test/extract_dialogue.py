import os
import json
import pandas as pd
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def extract_dialogue_data(folder_path, desired_languages):
    """
    Extracts dialogue data from JSON files in a specified folder for given languages.

    Args:
        folder_path (str): The path to the folder containing the JSON files.
        desired_languages (list): A list of language keys (strings) to extract.

    Returns:
        pandas.DataFrame: A DataFrame containing the extracted dialogue data.
    """
    all_data = []

    print(f"Processing files in folder: {folder_path}")

    for filename in os.listdir(folder_path):
        if filename.endswith(".json"):
            file_path = os.path.join(folder_path, filename)
            # print(f"  Processing file: {filename}") # Reduced console output

            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # Find the PSExternalMediaAsset object
                ps_asset = None
                for item in data:
                    if item.get("Type") == "PSExternalMediaAsset":
                        ps_asset = item
                        break

                if ps_asset and "Properties" in ps_asset and "LocalisedDialogueData" in ps_asset["Properties"]:
                    localised_data = ps_asset["Properties"]["LocalisedDialogueData"]

                    for entry in localised_data:
                        lang_key = entry.get("Key")
                        lang_value = entry.get("Value")

                        if lang_key and lang_value and lang_key in desired_languages:
                             # Extract properties from the specific language's Value dictionary
                            section = lang_value.get("Section", "N/A")
                            sub_section = lang_value.get("SubSection", "N/A")
                            sub_section_type = lang_value.get("SubSectionType", "N/A")
                            dialogue_context = lang_value.get("DialogueContext", "N/A")
                            dialogue_key = lang_value.get("DialogueKey", "N/A")
                            is_placeholder = lang_value.get("bIsPlaceholder", "N/A")
                            character_name = lang_value.get("CharacterName", "N/A")

                            # Extract and combine DisplayText from SubtitleLines
                            combined_display_text = ""
                            subtitle_lines = lang_value.get("SubtitleLines", [])
                            if subtitle_lines:
                                display_texts = [
                                    line.get("DisplayText", "")
                                    for line in subtitle_lines
                                    if line is not None and "DisplayText" in line # Ensure line is not None and has DisplayText key
                                ]
                                # Join non-empty display texts with a newline
                                combined_display_text = "\n".join(filter(None, display_texts))


                            all_data.append({
                                "Language": lang_key,
                                "DialogueKey": dialogue_key,
                                "Section": section,
                                "SubSection": sub_section,
                                "SubSectionType": sub_section_type,
                                "DialogueContext": dialogue_context,
                                "IsPlaceholder": is_placeholder,
                                "CharacterName": character_name,
                                "DisplayText": combined_display_text # Added DisplayText
                            })
                # else: # Reduced console output for files without the target asset
                    # print(f"    Warning: Could not find PSExternalMediaAsset or LocalisedDialogueData in {filename}")

            except json.JSONDecodeError:
                print(f"    Error: Could not decode JSON from {filename}")
            except Exception as e:
                print(f"    An unexpected error occurred while processing {filename}: {e}")

    return pd.DataFrame(all_data)

# Helper function to create valid Excel sheet names
def create_valid_sheet_name(name):
    """Cleans a string to be a valid Excel sheet name."""
    if pd.isna(name) or name is None or str(name).strip() == "": # Handle None, NaN, or empty strings
        return "No_SubSection"
    name = str(name) # Ensure it's a string
    # Replace invalid characters for Excel sheet names
    invalid_chars = r'[\\/:?*\[\]]'
    cleaned_name = re.sub(invalid_chars, '_', name)
    # Excel sheet names cannot start or end with an apostrophe
    cleaned_name = cleaned_name.strip("'")
     # Replace leading/trailing spaces and dots
    cleaned_name = cleaned_name.strip(' .')
    # Replace multiple underscores with a single one
    cleaned_name = re.sub(r'_{2,}', '_', cleaned_name)
    # Truncate to max 31 characters
    return cleaned_name[:31]


if __name__ == "__main__":
    folder_path = input("Enter the path to the folder containing the JSON files: ")

    if not os.path.isdir(folder_path):
        print(f"Error: Folder not found at '{folder_path}'")
    else:
        languages_input = input("Enter the languages to extract (comma-separated, e.g., English,German,Japanese): ")
        # Clean the input languages
        desired_languages = [lang.strip() for lang in languages_input.split(',') if lang.strip()]

        if not desired_languages:
            print("No valid languages specified. Exiting.")
        else:
            print(f"Attempting to extract data for languages: {', '.join(desired_languages)}")
            extracted_df = extract_dialogue_data(folder_path, desired_languages)

            if not extracted_df.empty:
                # --- Auto Save to Excel with language prefix and grouped sheets ---
                # Create a language string for the filename from the *requested* languages
                language_prefix = "_".join(desired_languages)
                # Replace any characters that might be invalid in filenames (optional but good practice)
                # Keep only alphanumeric characters and underscores for safety
                language_prefix_safe = "".join([c if c.isalnum() or c == '_' else '_' for c in language_prefix])
                # Remove consecutive underscores or leading/trailing ones resulting from cleaning
                language_prefix_safe = '_'.join(filter(None, language_prefix_safe.split('_')))
                # Ensure a trailing underscore if there's a prefix, unless it's the only character
                if language_prefix_safe and len(language_prefix_safe) > 0 and not language_prefix_safe.endswith('_'):
                     language_prefix_safe += '_'
                # Remove leading underscore if it exists
                if language_prefix_safe.startswith('_'):
                     language_prefix_safe = language_prefix_safe[1:]


                default_output_filename = f"{language_prefix_safe}extracted_dialogue.xlsx"
                 # If the prefix ends up empty after cleaning, just use the default name
                if not language_prefix_safe:
                     default_output_filename = "extracted_dialogue.xlsx"


                # Save the file in the current working directory where the script is run
                output_path = os.path.join(os.getcwd(), default_output_filename)

                try:
                    print(f"\nSaving data to '{output_path}' grouped by SubSection and applying formatting...")
                    # Create an Excel writer object
                    # You need to have the openpyxl engine installed: pip install openpyxl
                    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                        # Group the DataFrame by 'SubSection'
                        # Sorting by SubSection makes the sheet order predictable
                        extracted_df = extracted_df.sort_values(by='SubSection')
                        grouped = extracted_df.groupby('SubSection')

                        # Write each group to a different sheet
                        for sub_section_name, group_df in grouped:
                            # Create a valid sheet name from the SubSection name
                            sheet_name = create_valid_sheet_name(sub_section_name)
                            # Ensure sheet name is not empty after cleaning
                            if not sheet_name:
                                sheet_name = "Unnamed_Section"

                            # Write the dataframe slice to the sheet
                            try:
                                group_df.to_excel(writer, sheet_name=sheet_name, index=False)

                                # --- Apply Formatting to the Sheet ---
                                worksheet = writer.sheets[sheet_name] # Get the openpyxl worksheet object

                                # 1. Freeze the first row
                                worksheet.freeze_panes = 'A2'

                                # 2. Define Alignment styles
                                center_aligned_text = Alignment(horizontal='center', vertical='center')
                                left_aligned_text = Alignment(horizontal='left', vertical='center', wrap_text=True) # Added wrap_text

                                # Find the column index for 'DisplayText' in this specific sheet's DataFrame
                                try:
                                     # Find the column index based on the header row text
                                     header = [cell.value for cell in worksheet[1]]
                                     display_text_col_idx = header.index('DisplayText') # Get 0-based index
                                except ValueError:
                                     # 'DisplayText' column not found in this sheet's header
                                     print(f"      Warning: 'DisplayText' column header not found in sheet '{sheet_name}'. Skipping specific alignment.")
                                     display_text_col_idx = -1 # Use -1 to indicate not found


                                # Iterate over all rows and apply alignment based on column
                                # Start from row 1 to include the header row for potential formatting
                                # Use iter_rows(min_row=1) to ensure headers are included in alignment loop
                                for row_index, row in enumerate(worksheet.iter_rows(min_row=1)):
                                    for col_index, cell in enumerate(row):
                                        if col_index == display_text_col_idx:
                                            cell.alignment = left_aligned_text
                                        else:
                                            cell.alignment = center_aligned_text

                                # 3. Autofit column width
                                # Calculate max length for each column
                                for col_idx_auto, col in enumerate(worksheet.columns):
                                    max_length = 0
                                    # Iterate through cells in the column
                                    for cell in col:
                                        try:
                                            # Check length of cell content, handle None/empty
                                            if cell.value is not None:
                                                 # Use str() to handle numbers/other types, and account for newlines
                                                 # Newlines make the cell taller, but don't affect width directly
                                                 # We only care about the longest *line* or the total character count for width
                                                 # Let's use the max length of any line within the cell or the total length
                                                 cell_string = str(cell.value)
                                                 current_max = max(len(line) for line in cell_string.split('\n')) if '\n' in cell_string else len(cell_string)
                                                 max_length = max(max_length, current_max)
                                            else:
                                                 cell_length = 0
                                            # Also consider header length for autofit
                                            # The header is in the first row (row 1)
                                            header_cell = worksheet.cell(row=1, column=col_idx_auto+1)
                                            if header_cell.value is not None:
                                                 max_length = max(max_length, len(str(header_cell.value)))


                                        except Exception as cell_e: # Catch any potential errors in cell value or str conversion
                                             # print(f"        Warning: Error getting length of cell in column {get_column_letter(col_idx_auto+1)}: {cell_e}") # Too noisy
                                             pass # Just skip problematic cells for length calculation


                                    # Apply width - rough estimate, might need tuning based on font/etc.
                                    # Excel column width units are based on the width of the digit 0
                                    # A common conversion is roughly char_length * 0.8 + 5 (for default font)
                                    # Or simply (max_length + padding) * factor. Let's use max_length * ~0.9 + base padding.
                                    adjusted_width = (max_length + 1) * 0.9 # Add padding, apply factor (tuned slightly)
                                    # Set a minimum width to ensure columns aren't too narrow
                                    min_width = 10 # Keep a reasonable minimum
                                    final_width = max(adjusted_width, min_width)

                                    # Set the width using the column letter
                                    column_letter = get_column_letter(col_idx_auto + 1)
                                    worksheet.column_dimensions[column_letter].width = final_width


                                # --- End Formatting ---

                            except Exception as sheet_e:
                                print(f"      Error writing and formatting sheet '{sheet_name}' for SubSection '{sub_section_name}': {sheet_e}")


                    print("Data successfully saved to Excel.")
                except ImportError:
                     print("\nError: openpyxl not found. Please install it using 'pip install openpyxl' to export to Excel.")
                except Exception as e:
                    print(f"\nError saving Excel file '{output_path}': {e}")
                # --- End Auto Save ---

            else:
                print("\nNo data extracted for the specified languages or JSON structure mismatch.")