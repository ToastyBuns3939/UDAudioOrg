import json
import pandas as pd
import os
import logging
from collections import defaultdict
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension # Import ColumnDimension

def format_file_size(size_in_bytes):
    """Converts bytes to a human-readable format (KB, MB, GB)."""
    if size_in_bytes is None:
        return "N/A"
    size_in_bytes = int(size_in_bytes) # Ensure it's an integer
    if size_in_bytes < 1024:
        return f"{size_in_bytes} Bytes"
    elif size_in_bytes < 1024 * 1024:
        return f"{size_in_bytes / 1024:.2f} KB"
    elif size_in_bytes < 1024 * 1024 * 1024:
        return f"{size_in_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_in_bytes / (1024 * 1024 * 1024):.2f} GB"


def convert_duplicates_json_to_excel(json_path="ps4_wem_analysis.json", excel_path="ps4_wem_analysis.xlsx"):
    """
    Converts the ps4_wem_analysis.json file into an Excel (.xlsx) file,
    with each category of files in a separate worksheet.
    Includes columns for Filename, Relative Path (aggregated), File Size, Duration,
    and All Categories.
    Sorts by 'Filename', applies wrap text and middle vertical alignment to all data cells,
    auto-sizes rows and columns (shared across sheets), and freezes the header row and column A.
    Adds filters to the header row.

    Args:
        json_path (str): The path to the input JSON file.
                         Defaults to 'ps4_wem_analysis.json'.
        excel_path (str): The path for the output Excel file.
                          Defaults to 'ps4_wem_analysis.xlsx'.
    """
    if not os.path.exists(json_path):
        logging.error(f"❌ Error: Input JSON file not found at {json_path}")
        print(f"Error: Input JSON file not found at {json_path}")
        return

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        logging.error(f"❌ Error decoding JSON from {json_path}: {e}")
        print(f"Error decoding JSON from {json_path}: {e}")
        return
    except Exception as e:
        logging.error(f"❌ An error occurred while reading {json_path}: {e}")
        print(f"An error occurred while reading {json_path}: {e}")
        return

    if not data:
        logging.info("✅ No file data found in JSON to write to Excel.")
        print("No file data found in JSON to write to Excel.")
        return

    try:
        # Use ExcelWriter to write to multiple sheets
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            all_sheets_dataframes = {} # Store DataFrames for calculating shared column widths

            for category, filenames_data in data.items():
                # Prepare data for pandas DataFrame for the current category
                excel_data_category = []
                for filename, file_info in filenames_data.items():
                    # Extract paths, all_categories, size, and duration from the file_info dictionary
                    paths = file_info.get("paths", [])
                    all_categories = file_info.get("all_categories", [])
                    file_size = file_info.get("size", 0) # Get size
                    duration = file_info.get("duration", "N/A") # Get duration (will be calculated value now)

                    # Convert the list of all categories to a comma-separated string for the Excel cell
                    all_categories_str = ", ".join(all_categories)

                    # Join all paths into a single string with line breaks
                    relative_paths_str = "\n".join(paths)

                    # Append a single row for this filename
                    excel_data_category.append({
                        "Filename": filename,
                        "Relative Path": relative_paths_str, # Use the aggregated string
                        "File Size": format_file_size(file_size), # Add formatted size
                        "Duration": duration, # Add duration (calculated value)
                        "All Categories": all_categories_str # Add the new column
                    })

                if excel_data_category:
                    # Create a pandas DataFrame for the current category
                    df_category = pd.DataFrame(excel_data_category)

                    # Sort by 'Filename'
                    df_category = df_category.sort_values(by="Filename", ascending=True)

                    # Sanitize category name for sheet name (Excel sheet names have limits)
                    sheet_name = category.replace(":", "_").replace("/", "_").replace("\\", "_")[:31] # Limit to 31 chars
                    df_category.to_excel(writer, index=False, sheet_name=sheet_name)
                    logging.info(f"  - Wrote category '{category}' to sheet '{sheet_name}' in {excel_path}")

                    all_sheets_dataframes[sheet_name] = df_category # Store DataFrame

            # --- Apply Formatting, Freezing, Auto-sizing, and Filters after all sheets are written ---
            if all_sheets_dataframes:
                # Calculate shared column widths across all sheets
                max_widths = defaultdict(int)
                for sheet_name, df in all_sheets_dataframes.items():
                     # Include header row width in calculation
                     for i, col in enumerate(df.columns):
                         max_widths[i] = max(max_widths[i], len(str(col))) # Check header width
                         for row_idx in range(len(df)):
                             cell_value = str(df.iloc[row_idx, i])
                             # Estimate width for wrapped text
                             max_widths[i] = max(max_widths[i], len(cell_value))


                # Apply formatting to each sheet
                for sheet_name, df in all_sheets_dataframes.items():
                    worksheet = writer.sheets[sheet_name]

                    # Apply Wrap Text and Middle Alignment to all data cells (skipping header)
                    for row in worksheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='center')

                    # Auto-size rows (skipping header)
                    for row_idx in range(2, worksheet.max_row + 1):
                         worksheet.row_dimensions[row_idx].auto_size = True

                    # Apply shared auto-size columns
                    for i, col in enumerate(df.columns):
                        column_letter = get_column_letter(i + 1) # +1 because openpyxl is 1-indexed
                        adjusted_width = (max_widths[i] + 2) * 1.2 # Add some padding
                        # Limit max width to avoid extremely wide columns
                        if adjusted_width > 100:
                            adjusted_width = 100
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                    # Freeze the header row and column A
                    worksheet.freeze_panes = 'B2'

                    # Add filters to the header row
                    # Determine the range of cells for the filter
                    filter_range = f"A1:{get_column_letter(len(df.columns))}{worksheet.max_row}"
                    worksheet.auto_filter.ref = filter_range

                    logging.debug(f"  - Applied formatting, auto-sizing, freezing, and filters in sheet '{sheet_name}'")

            else:
                 logging.info("  - No dataframes generated, skipping formatting.")


        logging.info(f"✅ Successfully converted {json_path} to {excel_path} with categories as sheets.")
        print(f"Successfully converted {json_path} to {excel_path} with categories as sheets.")


    except ImportError:
        logging.error("❌ Error: 'openpyxl' library not found. Please install it (`pip install openpyxl`) to write .xlsx files.")
        print("Error: 'openpyxl' library not found. Please install it (`pip install openpyxl`) to write .xlsx files.")
    except Exception as e:
        logging.error(f"❌ An error occurred while writing to Excel file {excel_path}: {e}")
        print(f"An error occurred while writing to Excel file {excel_path}: {e}")


def convert_mapping_json_to_excel(json_path="wem_mapping.json", excel_path="wem_mapping.xlsx", pc_wem_dir=None):
    """
    Converts the wem_mapping.json file into an Excel (.xlsx) file,
    with each top-level source JSON folder in a separate worksheet.
    Includes MediaPathName, DebugName, all Source JSONs (with line breaks),
    File Size, Duration, and a list of sheets the entry appears on.
    Sorts by 'DebugName', applies wrap text and middle vertical alignment to all data cells,
    auto-sizes rows and columns (shared across sheets), and freezes the header row and column A.
    Adds filters to the header row.
    Calculates duration by converting WEM to OGG if pc_wem_dir is provided.

    Args:
        json_path (str): The path to the input JSON file (wem_mapping.json).
                         Defaults to 'wem_mapping.json'.
        excel_path (str): The path for the output Excel file.
                          Defaults to 'wem_mapping.xlsx'.
        pc_wem_dir (str, optional): The path to the PC WEM directory (WwiseStaged).
                                    Required to get file sizes and duration. Defaults to None.
    """
    if not os.path.exists(json_path):
        logging.error(f"❌ Error: Input JSON file not found at {json_path}")
        print(f"Error: Input JSON file not found at {json_path}")
        return

    if pc_wem_dir is None:
        logging.warning("⚠️ PC WEM directory not provided. File Size and Duration will not be available in the PC mapping Excel.")
        print("Warning: PC WEM directory not provided. File Size and Duration will not be available in the PC mapping Excel.")
        # Proceed without file size/duration if directory is missing

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        logging.error(f"❌ Error decoding JSON from {json_path}: {e}")
        print(f"Error decoding JSON from {json_path}: {e}")
        return
    except Exception as e:
        logging.error(f"❌ An error occurred while reading {json_path}: {e}")
        print(f"An error occurred while reading {json_path}: {e}")
        return

    if not data:
        logging.info("✅ No mapping data found in JSON to write to Excel.")
        print("No mapping data found in JSON to write to Excel.")
        return

    # First, determine which sheet(s) each MediaPathName will appear on
    media_path_to_sheets = defaultdict(set)
    for media_path, map_data in data.items():
        source_jsons = map_data.get("SourceJsons", [])
        for source_json_path in source_jsons:
            # Extract the first directory and sanitize it for sheet name
            parts = source_json_path.replace("\\", "/").split("/")
            first_directory = parts[0] if parts else "Root"
            sanitized_sheet_name = first_directory.replace(":", "_").replace("/", "_").replace("\\", "_")[:31]
            media_path_to_sheets[media_path].add(sanitized_sheet_name)

    # --- Get File Size and Duration for the actual PC WEM files using ThreadPoolExecutor ---
    wem_files_to_process = []
    # Create a mapping from media_path to full wem file path for processing
    media_path_to_full_path = {}
    if pc_wem_dir:
        for media_path in data.keys():
             full_wem_path = os.path.normpath(os.path.join(pc_wem_dir, media_path))
             if os.path.exists(full_wem_path):
                 wem_files_to_process.append(full_wem_path)
                 media_path_to_full_path[media_path] = full_wem_path
             else:
                 logging.debug(f"  - PC WEM file not found for mapping entry: {full_wem_path}")

    file_info_results = {} # Store results {media_path: {"size": size, "duration": duration_string}}

    if wem_files_to_process and os.path.exists(WW2OGG_PATH) and os.path.exists(REVORB_PATH) and MUTAGEN_AVAILABLE:
        logging.info(f"Starting PC WEM analysis (size and duration) for {len(wem_files_to_process)} files...")
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_wem = {executor.submit(convert_wem_to_ogg_and_get_duration, wem_path): wem_path for wem_path in wem_files_to_process}

            for future in as_completed(future_to_wem):
                wem_path = future_to_wem[future]
                try:
                    duration = future.result()
                    # Find the original media_path from the full_wem_path
                    original_media_path = None
                    for mp, fp in media_path_to_full_path.items():
                         if fp == wem_path:
                             original_media_path = mp
                             break

                    if original_media_path:
                         file_size = os.path.getsize(wem_path) # Get size after confirming file exists
                         file_info_results[original_media_path] = {"size": file_size, "duration": duration}
                         logging.debug(f"Processed PC WEM {os.path.basename(wem_path)}: Size {file_size}, Duration {duration}")
                    else:
                         logging.warning(f"⚠️ Could not map processed WEM path back to original media_path: {wem_path}")


                except Exception as e:
                    logging.error(f"❌ Exception processing PC WEM {os.path.basename(wem_path)} for size/duration: {e}")
                    # Add placeholder if processing fails
                    original_media_path = None
                    for mp, fp in media_path_to_full_path.items():
                         if fp == wem_path:
                             original_media_path = mp
                             break
                    if original_media_path:
                         file_info_results[original_media_path] = {"size": 0, "duration": "N/A (Processing Error)"}


    else:
        logging.warning("⚠️ Skipping PC WEM size and duration calculation due to missing PC WEM directory, tools (ww2ogg.exe, revorb.exe) or mutagen library.")
        # Populate with default N/A values if processing is skipped
        for media_path in data.keys():
             file_info_results[media_path] = {"size": 0, "duration": "N/A (Tools/Library Missing)"}
    # ---------------------------------------------------------------------------------------------


    # Group data by the first directory in the Source JSON paths and aggregate Source JSONs
    grouped_data_aggregated = defaultdict(dict) # SheetName -> { (MediaPath, DebugName): { row_data } }
    for media_path, map_data in data.items():
        debug_name = map_data.get("DebugName", "")
        source_jsons = map_data.get("SourceJsons", [])

        # Get the sorted list of sheets this MediaPathName appears on
        appears_on_sheets_list = sorted(list(media_path_to_sheets[media_path]))
        # Convert list of sheets to a string with comma and space separator
        appears_on_sheets_str = ", ".join(appears_on_sheets_list)

        # Convert list of all source jsons to a string with line breaks for the 'Source JSON File' column
        all_source_jsons_str = "\n".join(source_jsons)

        # Get file size and duration from the processed results
        file_info = file_info_results.get(media_path, {"size": 0, "duration": "N/A (Not Processed)"})
        file_size = file_info["size"]
        duration = file_info["duration"]


        for source_json_path in source_jsons:
            # Extract the first directory from the relative path (for grouping)
            parts = source_json_path.replace("\\", "/").split("/")
            first_directory = parts[0] if parts else "Root"
            sheet_name = first_directory # Use the original first directory for grouping key

            # Use (media_path, debug_name) as the key to ensure uniqueness per sheet
            unique_key = (media_path, debug_name)

            # Add or update the entry for this unique key on the sheet
            # This ensures only one row per MediaPath/DebugName pair per sheet
            grouped_data_aggregated[sheet_name][unique_key] = {
                "MediaPathName (Wwise ID)": media_path,
                "DebugName": debug_name,
                "Source JSON File": all_source_jsons_str, # List ALL source JSONs here
                "File Size": format_file_size(file_size), # Add formatted size
                "Duration": duration, # Add duration
                "Group": appears_on_sheets_str # Renamed column to "Group"
            }

    try:
        # Use ExcelWriter to write to multiple sheets
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            all_sheets_dataframes = {} # Store DataFrames for calculating shared column widths

            # Sort sheet names alphabetically and process each group
            for sheet_name in sorted(grouped_data_aggregated.keys()):
                # Convert the dictionary of unique rows into a list of dictionaries for the DataFrame
                excel_data_sheet = list(grouped_data_aggregated[sheet_name].values())

                if excel_data_sheet:
                    df_sheet = pd.DataFrame(excel_data_sheet)

                    # Sort by 'DebugName'
                    df_sheet = df_sheet.sort_values(by="DebugName", ascending=True)

                    # Sanitize sheet name (for actual Excel sheet name)
                    sanitized_sheet_name = sheet_name.replace(":", "_").replace("/", "_").replace("\\", "_")[:31] # Limit to 31 chars
                    df_sheet.to_excel(writer, index=False, sheet_name=sanitized_sheet_name)
                    logging.info(f"  - Wrote data for source folder '{sheet_name}' to sheet '{sanitized_sheet_name}' in {excel_path}")

                    all_sheets_dataframes[sanitized_sheet_name] = df_sheet # Store DataFrame using sanitized name

            # --- Apply Formatting, Freezing, Auto-sizing, and Filters after all sheets are written ---
            if all_sheets_dataframes:
                # Calculate shared column widths across all sheets
                max_widths = defaultdict(int)
                # Iterate through the stored dataframes to find max widths
                for sheet_name, df in all_sheets_dataframes.items():
                     # Include header row width in calculation
                     for i, col in enumerate(df.columns):
                         max_widths[i] = max(max_widths[i], len(str(col))) # Check header width
                         for row_idx in range(len(df)):
                             cell_value = str(df.iloc[row_idx, i])
                             # Estimate width for wrapped text
                             max_widths[i] = max(max_widths[i], len(cell_value))


                # Apply formatting to each sheet
                for sheet_name, df in all_sheets_dataframes.items():
                    worksheet = writer.sheets[sheet_name]

                    # Apply Wrap Text and Middle Alignment to all data cells (skipping header)
                    for row in worksheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='center')

                    # Auto-size rows (skipping header)
                    for row_idx in range(2, worksheet.max_row + 1):
                         worksheet.row_dimensions[row_idx].auto_size = True

                    # Apply shared auto-size columns
                    for i, col in enumerate(df.columns):
                        column_letter = get_column_letter(i + 1) # +1 because openpyxl is 1-indexed
                        adjusted_width = (max_widths[i] + 2) * 1.2 # Add some padding
                        # Limit max width to avoid extremely wide columns
                        if adjusted_width > 100:
                            adjusted_width = 100
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                    # Freeze the header row and column A
                    worksheet.freeze_panes = 'B2'

                    # Add filters to the header row
                    # Determine the range of cells for the filter
                    filter_range = f"A1:{get_column_letter(len(df.columns))}{worksheet.max_row}"
                    worksheet.auto_filter.ref = filter_range

                    logging.debug(f"  - Applied formatting, auto-sizing, freezing, and filters in sheet '{sanitized_sheet_name}'")


            else:
                 logging.info("  - No dataframes generated, skipping formatting.")


        logging.info(f"✅ Successfully converted {json_path} to {excel_path} with source folders as sheets.")
        print(f"Successfully converted {json_path} to {excel_path} with source folders as sheets.")
        if pc_wem_dir is None:
             logging.warning("⚠️ File Size and Duration columns may be empty as PC WEM directory was not provided.")


    except ImportError:
        logging.error("❌ Error: 'openpyxl' library not found. Please install it (`pip install openpyxl`) to write .xlsx files.")
        print("Error: 'openpyxl' library not found. Please install it (`pip install openpyxl`) to write .xlsx files.")
    except Exception as e:
        logging.error(f"❌ An error occurred while writing to Excel file {excel_path}: {e}")
        print(f"An error occurred while writing to Excel file {excel_path}: {e}")


if __name__ == "__main__":
    # Basic logging setup for the script
    logging.basicConfig(
        level=logging.INFO,
        format='[%(levelname)s] %(message)s',
        handlers=[logging.StreamHandler()]
    )

    # Example usage (can be run directly for testing)
    # convert_duplicates_json_to_excel()
    # convert_mapping_json_to_excel() # Note: This will not include file size/duration without pc_wem_dir
    pass # Avoid running conversion automatically when imported
